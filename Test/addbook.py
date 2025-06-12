import time
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import StaleElementReferenceException
from datetime import datetime

# === CẤU HÌNH ===
EXCEL_FILE = "testcase.xlsx"
SHEET_NAME = "ThemSanPham"
BASE_URL = "https://localhost:44329/Admin/themvasua"

# === DRIVER ===
driver = webdriver.Chrome()
driver.maximize_window()

# Đăng nhập
driver.get("https://localhost:44329/Login")
driver.find_element(By.ID, "txtUsername").send_keys("admin")
driver.find_element(By.ID, "txtPassword").send_keys("123")
Select(driver.find_element(By.ID, "ddlRole")).select_by_value("1")
driver.find_element(By.ID, "btnLogin").click()
time.sleep(2)

# Excel
wb = openpyxl.load_workbook(EXCEL_FILE)
sheet = wb[SHEET_NAME]
KET_QUA_COL = 9

for row in range(2, sheet.max_row + 1):
    errors = []
    test_id = sheet.cell(row=row, column=1).value
    ten = sheet.cell(row=row, column=2).value or ""
    danh_muc = sheet.cell(row=row, column=3).value or ""
    mo_ta = sheet.cell(row=row, column=4).value or ""
    gia = str(sheet.cell(row=row, column=5).value or "")
    so_luong = str(sheet.cell(row=row, column=6).value or "")
    ngay_xuat_ban_goc = sheet.cell(row=row, column=7).value or ""

    # === Kiểm tra riêng cho TCI_1: tất cả trường rỗng ===
    if test_id == "TCI_1":
        driver.get(BASE_URL)
        time.sleep(1)

        driver.find_element(By.ID, "MainContent_txtTenSach").clear()
        driver.find_element(By.ID, "MainContent_txtMoTa").clear()
        driver.find_element(By.ID, "MainContent_txtGia").clear()
        driver.find_element(By.ID, "MainContent_txtSoLuongTon").clear()
        driver.find_element(By.ID, "MainContent_txtNgayXuatBan").clear()
        Select(driver.find_element(By.ID, "MainContent_ddlDanhMuc")).select_by_index(0)

        driver.find_element(By.ID, "MainContent_btnSave").click()
        time.sleep(1)

        msg = driver.find_element(By.ID, "MainContent_lblMessage").text.strip()
        if "tồn tại" in msg.lower():
            ket_qua = f"Lỗi từ server: ⚠ tên sách đã tồn tại. vui lòng nhập tên khác."
        elif msg:
            ket_qua = f"Lỗi từ server: {msg}"
        else:
            ket_qua = "❌ Không hiển thị thông báo lỗi."

        sheet.cell(row=row, column=KET_QUA_COL).value = ket_qua
        continue  # Bỏ qua bước xử lý chung

    # Chuẩn hóa ngày xuất bản
    ngay_xuat_ban = ""
    try:
        if isinstance(ngay_xuat_ban_goc, datetime):
            ngay_xuat_ban = ngay_xuat_ban_goc.strftime("%Y-%m-%d")
        else:
            try:
                dt = datetime.strptime(str(ngay_xuat_ban_goc), "%d/%m/%Y")
            except:
                dt = datetime.strptime(str(ngay_xuat_ban_goc), "%Y-%m-%d")
            ngay_xuat_ban = dt.strftime("%Y-%m-%d")
    except:
        errors.append("⚠ Ngày xuất bản không đúng định dạng yyyy-MM-dd.")

    # Mở form
    driver.get(BASE_URL)
    time.sleep(1)

    # Nhập dữ liệu
    driver.find_element(By.ID, "MainContent_txtTenSach").clear()
    driver.find_element(By.ID, "MainContent_txtTenSach").send_keys(ten)

    driver.find_element(By.ID, "MainContent_txtMoTa").clear()
    driver.find_element(By.ID, "MainContent_txtMoTa").send_keys(mo_ta)

    driver.find_element(By.ID, "MainContent_txtGia").clear()
    driver.find_element(By.ID, "MainContent_txtGia").send_keys(gia)

    driver.find_element(By.ID, "MainContent_txtSoLuongTon").clear()
    driver.find_element(By.ID, "MainContent_txtSoLuongTon").send_keys(so_luong)

    driver.find_element(By.ID, "MainContent_txtNgayXuatBan").clear()
    driver.find_element(By.ID, "MainContent_txtNgayXuatBan").send_keys(ngay_xuat_ban)

    # Danh mục
    try:
        dropdown = Select(driver.find_element(By.ID, "MainContent_ddlDanhMuc"))
        found = False
        for option in dropdown.options:
            if option.text.strip().lower() == danh_muc.strip().lower():
                dropdown.select_by_visible_text(option.text)
                found = True
                break
        if not found:
            dropdown.select_by_index(0)
    except:
        errors.append("⚠ Không tìm thấy dropdown danh mục.")

    # Nhấn lưu
    driver.find_element(By.ID, "MainContent_btnSave").click()
    time.sleep(1)

    # Kiểm tra client-side
    if ten.strip() == "":
        errors.append("⚠ Tên sách không được để trống.")
    if mo_ta.strip() == "":
        errors.append("⚠ Mô tả không được để trống.")
    try:
        float(gia)
    except:
        errors.append("⚠ Giá không hợp lệ.")
    try:
        int(so_luong)
    except:
        errors.append("⚠ Số lượng không hợp lệ.")
    try:
        dropdown = Select(driver.find_element(By.ID, "MainContent_ddlDanhMuc"))
        if dropdown.first_selected_option.text.strip().startswith("--"):
            errors.append("⚠ Chưa chọn danh mục.")
    except StaleElementReferenceException:
        pass

    # Kết quả
    if not errors:
        msg = driver.find_element(By.ID, "MainContent_lblMessage").text.strip().lower()
        if "thành công" in msg:
            ket_qua = "Thêm thành công"
        else:
            ket_qua = f"Lỗi từ server: {msg}"
    else:
        ket_qua = ", ".join(errors)

    # Ghi lại
    sheet.cell(row=row, column=KET_QUA_COL).value = ket_qua

# Lưu kết quả
wb.save(EXCEL_FILE)
driver.quit()
print("✅ Hoàn tất kiểm thử. Kết quả đã ghi vào Excel.")
